import openpyxl
from openpyxl import Workbook
from pathlib import Path
import copy as copymod
from .base import copy_sheet, copy_data_sheet, fill_common_fields

TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
PRIMARY_TEMPLATES_DIR = TEMPLATES_DIR / "Primary_ITCs"
PROTECTION_SECONDARY_TEMPLATES_DIR = TEMPLATES_DIR / "Protection_Secondary_ITCs"

TEMPLATES = {
    "circuit_breaker_generic": PRIMARY_TEMPLATES_DIR / "Circuit_Breaker.xlsx",
    "current_transformer": PRIMARY_TEMPLATES_DIR / "Current_Transformer.xlsx",
    "voltage_transformer": PRIMARY_TEMPLATES_DIR / "Voltage_Transformer.xlsx",
    "neutral_ct": PRIMARY_TEMPLATES_DIR / "Neutral_CT.xlsx",
    "isolator": PRIMARY_TEMPLATES_DIR / "Isolator.xlsx",
    "earth_switch": PRIMARY_TEMPLATES_DIR / "Earth_Switch.xlsx",
    "ows": PRIMARY_TEMPLATES_DIR / "OWS.xlsx",
    "net": PRIMARY_TEMPLATES_DIR / "NET.xlsx",
    "power_transformer": PRIMARY_TEMPLATES_DIR / "Power_Transformer.xlsx",
    "surge_arrestor": PRIMARY_TEMPLATES_DIR / "Surge_Arrestor.xlsx",
    "sel_751_feeder_relay": PROTECTION_SECONDARY_TEMPLATES_DIR / "SEL_751_Feeder_Relay.xlsx",
    "ac_board": PRIMARY_TEMPLATES_DIR / "AC_Board.xlsx",
    "aux_tf": PRIMARY_TEMPLATES_DIR / "Aux_TF.xlsx",
    "bess_pcs": PRIMARY_TEMPLATES_DIR / "BESS_PCS.xlsx",
    "dc_panel": PRIMARY_TEMPLATES_DIR / "DC_Panel.xlsx",
    "feeder_panel": PRIMARY_TEMPLATES_DIR / "Feeder_Panel.xlsx",
}

LABELS = {
    "circuit_breaker_generic": "Circuit Breaker",
    "current_transformer": "Current Transformer",
    "voltage_transformer": "Voltage Transformer",
    "neutral_ct": "Neutral CT",
    "isolator": "Isolator",
    "earth_switch": "Earth Switch",
    "ows": "OWS",
    "net": "NET",
    "power_transformer": "Power Transformer",
    "surge_arrestor": "Surge Arrestor",
    "sel_751_feeder_relay": "SEL 751 Feeder Relay",
    "ac_board": "AC Board",
    "aux_tf": "Auxiliary Transformer",
    "bess_pcs": "BESS PCS",
    "dc_panel": "DC Panel",
    "feeder_panel": "Feeder Panel",
}


def load_as_xlsx(template_path: Path) -> openpyxl.Workbook:
    """Load a template — if it's an xlsm, strip VBA by copying into a fresh Workbook."""
    if str(template_path).endswith(".xlsm"):
        wb_orig = openpyxl.load_workbook(template_path, keep_vba=True)
        ws_src = wb_orig[wb_orig.sheetnames[0]]

        wb_clean = Workbook()
        ws_dst = wb_clean.active
        ws_dst.title = ws_src.title

        for key, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[key].width = dim.width
        for key, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[key].height = dim.height

        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copymod.copy(cell.font)
                    new_cell.border = copymod.copy(cell.border)
                    new_cell.fill = copymod.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copymod.copy(cell.protection)
                    new_cell.alignment = copymod.copy(cell.alignment)

        for merged in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged))

        return wb_clean
    else:
        return openpyxl.load_workbook(template_path)


def generate(equipment_type: str, data: dict, panel_numbers: list[str], output_path: Path):
    template = TEMPLATES[equipment_type]
    label = LABELS[equipment_type]

    wb_orig = load_as_xlsx(template)
    ws_src = wb_orig[wb_orig.sheetnames[0]]

    wb_new = Workbook()
    wb_new.remove(wb_new.active)
    copy_data_sheet(wb_orig, wb_new)

    for panel_num in panel_numbers:
        title = f"{label} {panel_num} SAT"
        ws = copy_sheet(ws_src, wb_new, panel_num)
        fill_common_fields(ws, data, panel_num, title)

    wb_new.save(output_path)
