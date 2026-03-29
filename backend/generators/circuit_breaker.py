import openpyxl
from pathlib import Path
from .base import copy_sheet, copy_data_sheet, fill_common_fields


def generate(data: dict, panel_numbers: list[str], output_path: Path, template_path: Path):
    wb_orig = openpyxl.load_workbook(template_path)
    ws_src = wb_orig[wb_orig.sheetnames[0]]

    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)
    copy_data_sheet(wb_orig, wb_new)

    for panel_num in panel_numbers:
        title = f"Circuit Breaker {panel_num} SAT"
        ws = copy_sheet(ws_src, wb_new, panel_num)
        fill_common_fields(ws, data, panel_num, title)

    wb_new.save(output_path)
