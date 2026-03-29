import copy
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


def copy_sheet(ws_src, wb_dst, sheet_name):
    ws_dst = wb_dst.create_sheet(sheet_name)

    for key, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[key].width = dim.width
    for key, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[key].height = dim.height

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    for merged in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged))

    # A4 landscape, fit all columns to one page wide
    ws_dst.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )
    ws_dst.page_setup.paperSize = ws_dst.PAPERSIZE_A4
    ws_dst.page_setup.orientation = ws_dst.ORIENTATION_LANDSCAPE
    ws_dst.page_setup.fitToWidth = 1
    ws_dst.page_setup.fitToHeight = 0
    ws_dst.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )

    return ws_dst


def copy_data_sheet(wb_src, wb_dst):
    if "Data" in wb_src.sheetnames:
        data_ws_src = wb_src["Data"]
        data_ws_dst = wb_dst.create_sheet("Data")
        for row in data_ws_src.iter_rows():
            for cell in row:
                data_ws_dst[cell.coordinate].value = cell.value


def fill_common_fields(ws, data: dict, panel_num: str, title: str):
    """Fill in the standard header fields common to all ITC types."""
    ws["A1"] = title

    # CPP project details
    ws["G9"] = data.get("cpp_project_name", "")
    ws["G10"] = data.get("cpp_job_no", "")
    ws["G11"] = data.get("bay_name", "")
    ws["G12"] = panel_num

    # Personnel
    ws["G13"] = data.get("prepared_by_name", "")
    ws["G14"] = data.get("prepared_by_position", "")
    ws["G15"] = data.get("date")
    if ws["G15"].value:
        ws["G15"].number_format = "DD/MM/YYYY"

    ws["G16"] = data.get("checked_by_name", "")
    ws["G17"] = data.get("checked_by_position", "")
    ws["G18"] = data.get("checked_by_signature", "")
    ws["G19"] = data.get("date")
    if ws["G19"].value:
        ws["G19"].number_format = "DD/MM/YYYY"

    # Client details
    ws["T9"] = data.get("client_project_title", "")
    ws["T10"] = data.get("client_project_number", "")
    ws["T11"] = data.get("site_location", "")
    ws["T13"] = data.get("client_checked_by_name", "")
    ws["T14"] = data.get("client_checked_by_position", "")
